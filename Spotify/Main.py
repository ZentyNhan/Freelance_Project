########## SECTION: Library ##########
from ast import Try, keyword
from contextlib import nullcontext
from lib2to3.pgen2.literals import test
import os
import time
from pickle import FALSE, TRUE
from subprocess import check_output
from inspect import currentframe
from typing import List
from xml.etree.ElementTree import QName
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException,ElementNotInteractableException
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from datetime import datetime
from datetime import date
import requests
from termcolor import colored 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import json

#Excel Processing:
from openpyxl import Workbook 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

#pypyodbc:
import pypyodbc as odbc
import pyodbc
#Web server:
from http.server import HTTPServer
from server import Server

HOST_NAME = 'localhost'
PORT = 8000

class delay():
    MINI_DELAY    = 0.5 #second
    SOFT_DELAY    = 1   #second
    DELAY         = 3   #second
    HARD_DELAY    = 5   #second
    MASSIVE_DELAY = 10  #second
    SUPER_DELAY   = 30  #second
    WORLD_DELAY   = 60  #second

class Process(delay):
    #Attributes:
    whoer_url    = f'https://whoer.net/fr'
    Spotify_url  = f'https://accounts.spotify.com/vi-VN/login?continue=https%3A%2F%2Fopen.spotify.com%2F'
    Profile_url  = f'https://www.spotify.com/tr/account/profile/'

    Element_dict = {
        'Nation_Sel' : '''//body/div[@id='__next']/div[1]/div[1]/div[2]/div[2]/div[2]/div[1]/article[1]/section[1]/form[1]/div[1]/button[1]'''
    }

    Json_ls = ['Status_TEST','Username_TEST', 'ID_Premium_TEST']

    #Constructor:
    def __init__(self, user_, password_, familyURL_):
        self.user     = user_
        self.password = password_
        self.familyUR = familyURL_

    def accessSpotify(self, driver):
        driver.get(self.Spotify_url)
        driver.maximize_window()
        sleep(self.DELAY)
        driver.find_element(By.ID ,'login-username').send_keys(self.user)
        driver.find_element(By.ID ,'login-password').send_keys(self.password)
        driver.find_element(By.ID ,'login-button').click()
        #Return:
        return True
    
    def switchNation(self, driver):
        driver.get(self.Profile_url)
        sleep(self.DELAY)
        select = Select(driver.find_element(By.ID, 'country'))
        select.select_by_value('TR')
        driver.find_element(By.XPATH, self.Element_dict['Nation_Sel'] ).click()
        #Return:
        return True

    def joinPremium(self, driver):
        #Return:
        return self.Json_ls

    @classmethod
    def checkCurrentIP(cls, driver):
        driver.get(cls.whoer_url)


if __name__ == "__main__": 
    #Local var
    Email       = ['z.ntnhan19@gmail.com']
    PassW       = ['Nhan0334842024']
    familyURL   = 'ABC'
    ret_dict    = {}

    #SERVER:
    httpd = HTTPServer((HOST_NAME,PORT),Server)
    print(time.asctime(), "Start Server - %s:%s"%(HOST_NAME,PORT))
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print('interrupted!')
    httpd.server_close()
    print(time.asctime(),'Stop Server - %s:%s' %(HOST_NAME,PORT))

    ########## ANCHOR: DO NOT CHANGE ##########
    #Instance:
    DRIVER = webdriver.Chrome('chromedriver.exe')
    USER   = Process(Email[0], PassW[0], familyURL)

    #Method:
    # USER.accessSpotify(DRIVER)
    # USER.HARD_DELAY
    # USER.switchNation(DRIVER)
    # USER.HARD_DELAY

    #Send request to Database:
    Status, Username, Token, ID_Premium = USER.joinPremium(DRIVER)
    ret_dict['JSON'] = dict(ret_dict)
    ret_dict['JSON'] = {
        'Status'     : Status,
        'Username'   : Username,
        'ID_Premium' : ID_Premium
    }

    requests.post(url = '', json = ret_dict['JSON'])

    print(ret_dict)




























