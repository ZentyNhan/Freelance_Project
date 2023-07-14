########## SECTION: Library ##########
from ast import Try, keyword
from contextlib import nullcontext
from lib2to3.pgen2.literals import test
import os
from pickle import FALSE, TRUE
from subprocess import check_output
from inspect import currentframe
from typing import List
from xml.etree.ElementTree import QName
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager #Window flatform only
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException,ElementNotInteractableException
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from datetime import datetime
from datetime import date
import requests
from termcolor import colored 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import json
import string
import time
import random

#Excel Processing:
from openpyxl import Workbook 
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

#pypyodbc:
import pypyodbc as odbc
import pyodbc

#self lib
# import lib

#DJANGO:
import django

print(django.get_version())

# DRIVER_NAME = 'ODBC Driver 17 for SQL Server'
# SERVER_NAME = 'DESKTOP-RCFES7K\SQLEXPRESS'
# DATABASE_NAME = 'SpotifyUpgrade_DB'

# connection = f'''
#         DRIVER={DRIVER_NAME};
#         SERVER={SERVER_NAME};
#         DATABASE={DATABASE_NAME};
#         UID=Thanhnhan19;
#         PWD=Nhan0334842024;
# '''
# #connect to database:
# conn = pyodbc.connect(connection)

# #use cursor:
# cursor = conn.cursor()

#Extract data from
#Local var
# Email = ['z.ntnhan19@gmail.com']
# PassW = ['Nhan0334842024']

# # for row in cursor.execute('select * from Customer_Infor'):
# #         Email.append(row.EmailAddress)
# #         PassW.append(row.EmailPW)

# print('Email: ',Email[0])
# print('PassW: ',PassW[0])

# driver = webdriver.Chrome('chromedriver.exe')
# url = 'https://whoer.net/fr'
# #Get link:
# driver.get(url)

# url = f'https://accounts.spotify.com/vi-VN/login?continue=https%3A%2F%2Fopen.spotify.com%2F'
# #Get link:
# driver.get(url)

# driver.maximize_window()
# sleep(3)
# driver.find_element(By.ID ,'login-username').send_keys(Email[0])
# driver.find_element(By.ID ,'login-password').send_keys(PassW[0])
# driver.find_element(By.ID ,'login-button').click()
# sleep(3)
# driver.get('https://www.spotify.com/tr/account/profile/')
# sleep(3)
# select = Select(driver.find_element(By.ID, 'country'))
# select.select_by_value('TR')

# driver.find_element(By.XPATH, '''//body/div[@id='__next']/div[1]/div[1]/div[2]/div[2]/div[2]/div[1]/article[1]/section[1]/form[1]/div[1]/button[1]''').click()


# input('Press any key to break')
# print("Press any key to break")




# familyURL   = 'https://www.spotify.com/vn-vi/family/join/invite/C3Cx803CX968Ya1/'

# list = familyURL.split('/')
# Index  = list.index('invite') + 1

# print(list[Index])
# a = [1,2,3]
# b = [*a,4,5,6]
# c = [4,5,6]

# c.extend(a) 
# print(c) # [1,2,3,4,5,6]


# *  giống extend
# ** giống update

# a = {'1':1,'2':2,'3':3,'4':4}
# b = {**a,'5':5,'6':6}
# print(b)

# def HAHAHA(a: int,b : int) -> str:
#         return a + b


# print(type(HAHAHA(5,4)))
# print(HAHAHA(5,10))
# n = 'aaaaaaa'
# o = n.count('a')
# print(o)

# Address     = 'Binbirdirek,__Peykhane__Cd.__10/A,__34122__Fatih/İstanbul,__Türkiye'

# # Email       = str(sys.argv[1])
# # PassW       = str(sys.argv[2])
# # familyURL   = str(sys.argv[3])

# #String Handling:
# Address    = Address.replace('__',' ')
# print(Address)


# Email       = 'z.ntnhan19@gmail.com'
# PassW       = 'Hihihi123@'
# Spotify_url  = f'https://accounts.spotify.com/vi-VN/login?continue=https%3A%2F%2Fopen.spotify.com%2F'
# op = webdriver.ChromeOptions()
# DRIVER  = webdriver.Chrome(ChromeDriverManager().install(), options=op)
# DRIVER.get(Spotify_url)
# WebDriverWait(DRIVER, 20).until(EC.presence_of_element_located((By.ID ,'login-username'))).send_keys(Email)
# WebDriverWait(DRIVER, 20).until(EC.presence_of_element_located((By.ID ,'login-password'))).send_keys(PassW)
# WebDriverWait(DRIVER, 20).until(EC.presence_of_element_located((By.ID ,'login-button'))).click()

# sleep(3)

# class logging():
#     #Attributes:
#     dt_format = "%d-%m-%Y_%H.%M.%S"
#     dt = datetime.now().strftime(dt_format)

#     def __init__(self, curpath_):
#         self.curpath = curpath_
#         self.curlogpath = os.path.join(self.curpath, 'log')
#         self.create_log_folder()
#         self.create_log()

#     def create_log(self):
#         try:
#             f  = open(os.path.join(self.curlogpath, f'Selenium_log_{self.dt}.txt'), 'w')
#             f.writelines(f'[{self.dt}]: ########## START RECORDING LOG ##########')
#         except Exception as e:
#             print(f'Failed: {e}')

#     def create_log_folder(self):
#         try:
#             os.mkdir(self.curlogpath)
#         except FileExistsError:
#             pass
#         except Exception as e:
#             print(f'Failed: {e}')


# print(os.getcwd())
# log = logging(os.getcwd())

WB = Workbook()
WS = WB.active
list = []
def generateVC(size=8, chars=string.ascii_uppercase + string.digits):
        return 'SF' + ''.join(random.choice(chars) for _ in range(size))

for i in range(26):
    # str =  f'https://www.spotify.com/vn-vi/family/join/invite/{generateVC()}/'
    str =  f'{generateVC()}'
    list.append(str)
    
print(list)


